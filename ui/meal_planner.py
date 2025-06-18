#ui/meal_planner.py
import os
import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter.simpledialog import askstring
from tkinter.filedialog import asksaveasfilename
import webbrowser
from fpdf import FPDF
from openpyxl import Workbook

from services.storage import get_mealplan, save_mealplan
from services.recipes import get_recipe_info
from collections import defaultdict
from data.ingredient_map import normalization_map

import re


class MealPlannerPanel:
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.mealplan = [m for m in get_mealplan() if isinstance(m, dict) and "day" in m and "meal" in m and "title" in m]
        self.entries = {}
        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.meals = ["Breakfast", "Lunch", "Dinner"]
        self.setup_ui()

    def setup_ui(self):
        for widget in self.parent.winfo_children():
            widget.destroy()

        ttk.Label(self.parent, text="📅 " + self.app.tr("weekly_meal_planner"), font=("Segoe UI", 18, "bold")).pack(pady=10)

        wrapper = ttk.Frame(self.parent)
        wrapper.pack(fill="both", expand=True, padx=10, pady=5)

        canvas = tk.Canvas(wrapper)
        scrollbar_y = ttk.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
        scrollbar_x = ttk.Scrollbar(self.parent, orient="horizontal", command=canvas.xview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(fill="x")

        grid_frame = scrollable_frame

        ttk.Label(grid_frame, text="").grid(row=0, column=0, padx=10, pady=5)
        for col, day in enumerate(self.days, start=1):
            ttk.Label(grid_frame, text=day, font=("Segoe UI", 11, "bold")).grid(row=0, column=col, padx=5, pady=5)

        for row, meal in enumerate(self.meals, start=1):
            ttk.Label(grid_frame, text=meal, font=("Segoe UI", 11, "bold")).grid(row=row, column=0, padx=5, pady=5)
            for col, day in enumerate(self.days, start=1):
                matched = next((m for m in self.mealplan if m.get("day") == day and m.get("meal") == meal), None)
                if matched:
                    title = matched.get("title", "")
                    url = matched.get("sourceUrl", "")
                    recipe_id = matched.get("id")
                    label = ttk.Label(grid_frame, text=title, cursor="hand2", wraplength=140, justify="left")
                    label.grid(row=row, column=col, padx=4, pady=4, sticky="w")
                    label.bind("<Button-1>", lambda e, rid=recipe_id, url=url: self.open_recipe(rid, url))
                    self.entries[(day, meal)] = label
                else:
                    entry = ttk.Entry(grid_frame, width=22)
                    entry.grid(row=row, column=col, padx=4, pady=4)
                    self.entries[(day, meal)] = entry
                    entry.bind("<FocusOut>", lambda e, d=day, m=meal, box=entry: self.save_meal(d, m, box.get()))
                    ttk.Label(grid_frame, text=app.tr("ingredient_hint"), foreground="gray").grid(row=row+1, column=col, sticky="w", padx=4)


        btn_frame = ttk.Frame(self.parent)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text=self.app.tr("save_all"), bootstyle="success", command=self.save_all).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="➕ Add Meal", bootstyle="primary", command=self.prompt_add_meal).pack(side="left", padx=5)
        ttk.Button(btn_frame, text=self.app.tr("shopping_list"), bootstyle="info", command=self.show_shopping_list).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="❌ Clear Planner", bootstyle="danger", command=self.clear_planner).pack(side="left", padx=5)

    def open_recipe(self, rid, url):
        if not hasattr(self.app, 'open_recipe'):
            print('[MealPlanner] ERROR: app has no open_recipe method')
            return
        full = get_recipe_info(rid)
        if full:
            for entry in self.mealplan:
                if entry.get("id") == rid:
                    entry["ingredients"] = [i["original"] for i in full.get("extendedIngredients", []) if "original" in i]
                    break
            save_mealplan(self.mealplan)
            self.app.open_recipe(full, show_close=True)
        elif url:
            webbrowser.open(url)

    def save_meal(self, day, meal, title):
        if not title:
            return
        for entry in self.mealplan:
            if entry.get("day") == day and entry.get("meal") == meal:
                entry["title"] = title
                break
        else:
            self.mealplan.append({"day": day, "meal": meal, "title": title})
        save_mealplan(self.mealplan)

    def save_all(self):
        new_plan = []
        for (day, meal), widget in self.entries.items():
            if isinstance(widget, ttk.Entry):
                title = widget.get().strip()
                if title:
                    new_plan.append({"day": day, "meal": meal, "title": title})
        self.mealplan = new_plan
        save_mealplan(self.mealplan)
        tk.messagebox.showinfo("Saved", "Meal plan saved successfully!")

    def prompt_add_meal(self):
        day = askstring("Choose Day", "Enter day (e.g., Monday):")
        if not day or day not in self.days:
            tk.messagebox.showerror(self.app.tr("error"), "Invalid day entered.")
            return
        meal = askstring("Choose Meal", "Enter time (Breakfast, Lunch, Dinner):")
        if not meal or meal not in self.meals:
            tk.messagebox.showerror(self.app.tr("error"), "Invalid meal time.")
            return
        title = askstring("Meal Name", f"What recipe to assign to {day} {meal}?")
        if not title:
            return
        entry = self.entries.get((day, meal))
        if isinstance(entry, ttk.Entry):
            entry.delete(0, "end")
            entry.insert(0, title)
            self.save_meal(day, meal, title)

    def show_shopping_list(self):
        ingredients = []
        for entry in self.mealplan:
            if "ingredients" in entry:
                ingredients.extend(entry["ingredients"])

        if not ingredients:
            tk.messagebox.showinfo(self.app.tr("no_ingredients"), "No recipes with ingredients found. Try clicking recipe titles first.")
            return

        grouped = self.simplify_ingredients(ingredients)

        popup = tk.Toplevel(self.parent)
        popup.title("\U0001F6D2 Shopping List")
        popup.geometry("450x600")
        popup.resizable(True, True)

        ttk.Label(popup, text="Shopping List", font=("Segoe UI", 14, "bold")).pack(pady=(10, 0))

        text = tk.Text(popup, wrap="word", bg="#f0f0f0", font=("Segoe UI", 10))
        for base, originals in grouped.items():
            text.insert("end", f"{base}:\n")
            for item in originals:
                text.insert("end", f"  - {item}\n")
            text.insert("end", "\n")
        text.config(state="disabled")
        text.pack(fill="both", expand=True, padx=10, pady=(10, 0))

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill="x", pady=(10, 15))

        ttk.Button(
            btn_frame,
            text=self.app.tr("export"),
            bootstyle="secondary",
            command=lambda: self.export_shopping_list(grouped)
        ).pack()

    def export_shopping_list(self, grouped):
        filetypes = [("Text file", "*.txt"), ("PDF file", "*.pdf"), ("CSV file", "*.csv"), ("Excel file", "*.xlsx")]
        path = asksaveasfilename(defaultextension=".txt", filetypes=filetypes, title="Save Shopping List As")
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".txt":
                with open(path, "w", encoding="utf-8") as f:
                    for base, items in grouped.items():
                        f.write(f"{base}:\n")
                        for item in items:
                            f.write(f"  - {item}\n")
                        f.write("\n")
            elif ext == ".pdf":
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                pdf.cell(200, 10, txt="Shopping List", ln=1, align="C")
                for base, items in grouped.items():
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(200, 8, txt=f"{base}:", ln=1)
                    pdf.set_font("Arial", "", 12)
                    for item in items:
                        pdf.cell(200, 8, txt=f"  - {item}", ln=1)
                    pdf.cell(200, 5, ln=1)
                pdf.output(path)
            elif ext == ".csv":
                with open(path, "w", encoding="utf-8") as f:
                    f.write("Group,Item\n")
                    for base, items in grouped.items():
                        for item in items:
                            f.write(f"{base},{item}\n")
            elif ext == ".xlsx":
                wb = Workbook()
                ws = wb.active
                ws.title = "Shopping List"
                ws.append(["Group", "Item"])
                for base, items in grouped.items():
                    for item in items:
                        ws.append([base, item])
                wb.save(path)
            else:
                tk.messagebox.showerror(self.app.tr("invalid_format"), "Unsupported file type.")
                return
            tk.messagebox.showinfo(self.app.tr("exported"), f"Shopping list exported to:\n{path}")
        except Exception as e:
            tk.messagebox.showerror(self.app.tr("export_failed"), f"Could not export shopping list:\n{e}")

    def simplify_ingredients(self, raw_items):
        grouped = defaultdict(list)

        unit_pattern = r'\b(\d+\/\d+|\d+\.\d+|\d+)\s*(cup|cups|tbsp|tsp|teaspoon|tablespoon|package|slice|pieces|oz|ml|g|kg|pound|lb|clove|dash|pinch|quart|liter|stick|can|pkg|bunch|handful|boxes)?\b'
        descriptor_pattern = r'\b(chopped|sliced|minced|diced|fresh|optional|to taste|large|small|medium|extra|firm|soft|finely|roughly|thinly|ground|crushed|real|size|rinsed|cut|high grade|korean|baby|cooked|uncooked|prepared|regular|julienned|into 1inch cubes|available at asian markets)\b'

        for item in raw_items:
            original = item.strip()
            cleaned = original.lower()
            cleaned = re.sub(r'\(.*?\)', '', cleaned)                      # Remove parenthesis
            cleaned = re.sub(unit_pattern, '', cleaned)                   # Remove measurements
            cleaned = re.sub(descriptor_pattern, '', cleaned)            # Remove descriptors
            cleaned = re.sub(r'[^\w\s]', '', cleaned)                    # Remove punctuation
            cleaned = re.sub(r'\s+', ' ', cleaned).strip()               # Normalize spaces

            # Apply normalization map (exact match or fallback to cleaned)
            normalized = normalization_map.get(cleaned, cleaned)

            grouped[normalized].append(original)

        return dict(sorted(grouped.items()))


    def clear_planner(self):
        self.mealplan = []
        save_mealplan(self.mealplan)
        self.setup_ui()


def show_meal_planner(app):
    app.clear_main()
    MealPlannerPanel(app.main, app)
